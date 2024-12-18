import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Step 1: Read the CSV file
csv_file = 'cleaned_agents.csv'  # Replace with your CSV file path
data = pd.read_csv(csv_file)

# Rename and process columns
data = data.rename(columns={
    "rover-img-lazy": "Full Name",  # Original column with full names
    "phone": "Cell",
    "email": "e-mail"
})

data = data.fillna("N/A").replace("", "N/A")

# Split "Full Name" into "First" and "Last"
if "Full Name" in data.columns:
    data[['First', 'Last']] = data['Full Name'].str.split(' ', n=1, expand=True)
else:
    data['First'] = ''
    data['Last'] = ''

# Add missing "Company" column (fill with empty strings for all rows)
if "Company" not in data.columns:
    data['Company'] = ''  # Fill the entire column with blank values

# Add the new column "REALTOR/Lender/Commercial REALTOR/Commercial Lender" (fill with empty strings)
data['REALTOR/Lender/Commercial REALTOR/Commercial Lender'] = ''

# Reorder columns to match the desired format
desired_columns = [
    "Last", 
    "First", 
    "Company", 
    "REALTOR/Lender/Commercial REALTOR/Commercial Lender", 
    "Cell", 
    "e-mail"
]
data = data[desired_columns]

# Step 2: Save the data to an Excel file
excel_file = 'formatted_agents.xlsx'  # Specify the output Excel file
data.to_excel(excel_file, index=False, sheet_name='Sheet1', startrow=2)  # Start at row 3 for title

# Step 3: Apply formatting
wb = load_workbook(excel_file)
sheet = wb.active

# Add the title "Customer List 2024" in the first row
sheet['A1'] = "Customer List 2024"
sheet['A1'].font = Font(bold=True, size=14)
sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
sheet.merge_cells('A1:F1')  # Merge across all columns

# Format the header row (row 3 now)
header_font = Font(bold=True)
for col in range(1, sheet.max_column + 1):
    cell = sheet.cell(row=3, column=col)
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Adjust column widths for better readability
column_widths = {
    "A": 15,  # Last
    "B": 15,  # First
    "C": 30,  # Company
    "D": 50,  # REALTOR/Lender/Commercial REALTOR/Commercial Lender
    "E": 15,  # Cell
    "F": 30   # e-mail
}
for col, width in column_widths.items():
    sheet.column_dimensions[col].width = width

# Add borders to all cells for clean formatting
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.border = thin_border

# Save the updated Excel file
wb.save(excel_file)

print(f"Excel file '{excel_file}' formatted to match the desired style.")
